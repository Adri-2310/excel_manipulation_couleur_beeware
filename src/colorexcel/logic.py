"""
Logic module for Excel color manipulation.

This module provides functions to extract and apply colors from one Excel file to another
based on matching Implantation, Nom, and Prénom columns.
"""

import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import logging
import pandas as pd

logger = logging.getLogger(__name__)

# Constantes
ALPHA_PREFIX = "FF"


def apply_tint(rgb: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    """
    Apply tint to an RGB color.
    
    Args:
        rgb: Tuple of (R, G, B) values
        tint: Tint value (-1.0 to 1.0). Positive values lighten, negative darken.
        
    Returns:
        Tuple of (R, G, B) values with tint applied
    """
    if tint == 0:
        return rgb
    
    r, g, b = rgb
    
    if tint > 0:
        # Lighten: interpolate towards white (255, 255, 255)
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    else:
        # Darken: interpolate towards black (0, 0, 0)
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    
    return (r, g, b)


def hex_to_rvb(hex_color: str) -> tuple[int, int, int] | None:
    """
    Convert hex color code to RGB tuple.

    Args:
        hex_color: Hexadecimal color code (e.g., 'FF0000' or 'FFFF0000')

    Returns:
        Tuple of (R, G, B) values or None if invalid
    """
    if hex_color is None:
        return None
    if hex_color.startswith(ALPHA_PREFIX):
        hex_color = hex_color[2:]
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return (r, g, b)
    except ValueError:
        logger.warning("Code couleur hex invalide : %s", hex_color, exc_info=True)
        return None


def extract_theme_colors(file_path: str) -> dict:
    """
    Extract theme colors from Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary mapping theme indices to hex color codes
    """
    theme_colors: dict[int, str] = {}
    if not os.path.exists(file_path):
        logger.error("Fichier thème introuvable : %s", file_path)
        return theme_colors
    try:
        with zipfile.ZipFile(file_path, "r") as zip_ref:
            if "xl/theme/theme1.xml" not in zip_ref.namelist():
                logger.warning("Pas de thème dans le fichier : %s", file_path)
                return theme_colors
            with zip_ref.open("xl/theme/theme1.xml") as theme_file:
                tree = ET.parse(theme_file)
                root = tree.getroot()
                ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
                for color_scheme in root.findall("a:themeElements/a:clrScheme", ns):
                    for i, color in enumerate(color_scheme):
                        rgb = color.find("a:srgbClr", ns)
                        if rgb is not None:
                            theme_colors[i] = rgb.attrib["val"]
    except Exception:
        logger.error("Erreur lors de l'extraction des couleurs du thème", exc_info=True)
    return theme_colors


def _find_columns_by_header(sheet, headers_wanted):
    """
    Find column indices by header names.

    Args:
        sheet: openpyxl worksheet object
        headers_wanted: Dict {logical_name: [list of possible header texts]}

    Returns:
        Dict {logical_name: column_index (0-based)} or None if any column is missing
    """
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    indices = {}

    for logical_name, candidates in headers_wanted.items():
        idx = None
        for i, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_text = str(cell_value).strip().lower()
            if cell_text in candidates:
                idx = i
                break
        if idx is None:
            logger.error(
                "Colonne '%s' introuvable dans l'en‑tête : %s", logical_name, header_row
            )
            return None
        indices[logical_name] = idx

    return indices


def get_sheet_names(file_path: str) -> list:
    """
    Get list of sheet names from an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of sheet names
    """
    try:
        if not os.path.exists(file_path):
            logger.error("Fichier introuvable : %s", file_path)
            return []

        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        logger.info("Feuilles trouvées dans %s : %s", file_path, sheet_names)
        return sheet_names
    except Exception:
        logger.error(
            "Erreur lors de la lecture du fichier : %s", file_path, exc_info=True
        )
        return []


def get_implantation_colors(file_path: str, sheet_name: str) -> dict:
    """
    Extract colors from the source Excel file based on Implantation, Nom, Prénom columns.

    Args:
        file_path: Path to the source Excel file
        sheet_name: Name of the sheet to read from

    Returns:
        Dictionary mapping (implantation, nom, prenom) tuples to RGB color tuples
    """
    if not os.path.exists(file_path):
        logger.error(
            "Fichier source introuvable pour get_implantation_colors : %s", file_path
        )
        return {}

    try:
        theme_colors = extract_theme_colors(file_path)
        workbook = load_workbook(filename=file_path, data_only=True)
    except Exception:
        logger.error(
            "Erreur lors de l'ouverture du fichier source : %s",
            file_path,
            exc_info=True,
        )
        return {}

    try:
        if sheet_name not in workbook.sheetnames:
            logger.error(
                "Feuille source introuvable : %s dans %s", sheet_name, file_path
            )
            return {}

        sheet = workbook[sheet_name]

        # Recherche dynamique des colonnes
        col_indices = _find_columns_by_header(
            sheet,
            {
                "implantation": ["implantation"],
                "nom": ["nom"],
                "prenom": ["prénom", "prenom"],
            },
        )
        if col_indices is None:
            return {}

        idx_impl = col_indices["implantation"]
        idx_nom = col_indices["nom"]
        idx_prenom = col_indices["prenom"]

        data_colors = {}

        # Parcourir à partir de la 2e ligne (1 = en-tête)
        for row in sheet.iter_rows(min_row=2):
            implantation = row[idx_impl].value
            nom = row[idx_nom].value
            prenom = row[idx_prenom].value

            if implantation is None or nom is None or prenom is None:
                continue

            key = (implantation, nom, prenom)

            cell_impl = row[idx_impl]
            if cell_impl.fill and cell_impl.fill.fill_type != "none":
                bg_color = cell_impl.fill.fgColor
                rvb_color = None
                tint = getattr(bg_color, 'tint', 0.0) or 0.0  # Gérer le tint

                if bg_color.type == "rgb":
                    rvb_color = hex_to_rvb(bg_color.rgb)
                elif bg_color.type == "theme":
                    hex_color = theme_colors.get(bg_color.theme)
                    if hex_color:
                        rvb_color = hex_to_rvb(hex_color)
                
                # Appliquer le tint si présent
                if rvb_color and tint != 0:
                    rvb_color = apply_tint(rvb_color, tint)

                # Ignorer les couleurs noires ou nulles
                if rvb_color and rvb_color != (0, 0, 0):
                    data_colors[key] = rvb_color
                    logger.debug(f"Couleur extraite pour {key}: RGB{rvb_color} (type: {bg_color.type}, tint: {tint})")

        return data_colors
    except Exception:
        logger.error(
            "Erreur lors de l'extraction des couleurs d'implantation", exc_info=True
        )
        return {}
    finally:
        workbook.close()


def apply_colors_to_file2(
    file1_path: str, file1_sheet: str, file2_path: str, file2_sheet: str
) -> str | None:
    """
    Apply colors from source file to a copy of target file based on matching Implantation, Nom, Prénom.

    Args:
        file1_path: Path to the source Excel file (with colors)
        file1_sheet: Sheet name in source file
        file2_path: Path to the target Excel file (to apply colors to)
        file2_sheet: Sheet name in target file
        
    Returns:
        Path to the new colored file or None if error
    """
    if not os.path.exists(file2_path):
        logger.error("Fichier cible introuvable : %s", file2_path)
        return None

    # Créer un fichier temporaire pour le traitement
    import tempfile
    import shutil
    from pathlib import Path
    
    temp_dir = tempfile.gettempdir()
    original_path = Path(file2_path)
    temp_file = Path(temp_dir) / f"colorexcel_temp_{original_path.name}"
    
    shutil.copy2(file2_path, temp_file)
    logger.info(f"Fichier temporaire créé : {temp_file}")

    data_colors = get_implantation_colors(file1_path, file1_sheet)

    try:
        workbook = load_workbook(temp_file)
    except Exception:
        logger.error(
            "Erreur lors de l'ouverture du fichier cible : %s",
            temp_file,
            exc_info=True,
        )
        return None

    try:
        if file2_sheet not in workbook.sheetnames:
            logger.error(
                "Feuille cible introuvable : %s dans %s", file2_sheet, temp_file
            )
            return None

        sheet = workbook[file2_sheet]

        # Même logique : trouver les colonnes Implantation/Nom/Prénom dans le fichier 2
        col_indices = _find_columns_by_header(
            sheet,
            {
                "implantation": ["implantation"],
                "nom": ["nom"],
                "prenom": ["prénom", "prenom"],
            },
        )
        if col_indices is None:
            return None

        idx_impl = col_indices["implantation"]
        idx_nom = col_indices["nom"]
        idx_prenom = col_indices["prenom"]

        for row in sheet.iter_rows(min_row=2):
            implantation = row[idx_impl].value
            nom = row[idx_nom].value
            prenom = row[idx_prenom].value

            if implantation is None or nom is None or prenom is None:
                continue

            key = (implantation, nom, prenom)
            rvb_color = data_colors.get(key)

            if rvb_color:
                fill = PatternFill(
                    start_color=f"{rvb_color[0]:02X}{rvb_color[1]:02X}{rvb_color[2]:02X}",
                    end_color=f"{rvb_color[0]:02X}{rvb_color[1]:02X}{rvb_color[2]:02X}",
                    fill_type="solid",
                )
                # Appliquer la couleur sur toute la ligne
                for cell in row:
                    cell.fill = fill

        workbook.save(temp_file)
        logger.info("Couleurs appliquées avec succès au fichier temporaire : %s", temp_file)
        return str(temp_file)
    except Exception:
        logger.error(
            "Erreur lors de l'application des couleurs au fichier cible", exc_info=True
        )
        return None
    finally:
        workbook.close()
